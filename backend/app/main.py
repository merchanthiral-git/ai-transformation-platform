"""AI Transformation Platform — FastAPI Backend v4.0
Modular architecture: core routes in separate router files,
extended analytics in main.py.
"""

import base64
import os
import sys
import traceback
from io import BytesIO
from pathlib import Path

import numpy as np
import pandas as pd
from fastapi import FastAPI, UploadFile, File, HTTPException, Query, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse

from app.store import (
    store, sync_work_design, build_reconstruction, build_ai_priority_scores,
    build_skill_analysis, build_deconstruction_summary, build_workstream_breakdown,
    build_task_portfolio, build_reconstruction_rollup, build_redeployment_detail,
    build_skill_shift_matrix, build_transformation_insights,
    build_transformation_recommendations, build_work_dimensions,
    build_value_model, build_role_evolution, build_capacity_waterfall_data,
    build_auto_change_plan, compute_readiness_score, get_manager_candidates,
    enrich_work_design_defaults, build_operating_model_analysis,
)
from app.helpers import get_series, safe_value_counts, dataframe_to_excel_bytes, empty_bundle
from app.shared import _safe, _j, _f, _job_list_for_model
from app.models import ALLOWED_EXTENSIONS, MAX_FILE_SIZE_BYTES, MAX_FILE_SIZE_MB
from app.storage import storage as file_storage, make_unique_filename

# ── Router imports ──
from app.routes_auth import auth_router, project_router
from app.routes_overview import router as overview_router
from app.routes_diagnose import router as diagnose_router
from app.routes_design import router as design_router
from app.routes_simulate import router as simulate_router
from app.routes_mobilize import router as mobilize_router
from app.routes_export import router as export_router
from app.routes_job_arch import router as job_arch_router
from app.routes_ja_mapping import router as ja_mapping_router
from app.routes_wd import router as wd_router
from app.routes_mgr import router as mgr_router
from app.routes_scorecard import router as scorecard_router
from app.routes_diagnostic import router as diagnostic_router
from app.routes_synthesis import router as synthesis_router
from app.routes_org_chart import router as org_chart_router
from app.routes_ai import router as ai_router
from app.routes_analytics import router as analytics_router
from app.routes_tutorial import router as tutorial_router
from app.routes_export_ext import router as export_ext_router
try:
    from app.routes_skills import router as skills_router
except ImportError:
    skills_router = None
try:
    from app.routes_job_content import router as job_content_router
except ImportError:
    job_content_router = None
try:
    from app.routes_notes import router as notes_router
except ImportError:
    notes_router = None

app = FastAPI(title="AI Transformation Platform API", version="4.0")

# ── Dev error handler ──
@app.exception_handler(Exception)
async def global_exception_handler(request: Request, exc: Exception):
    tb = traceback.format_exc()
    print(f"[500 ERROR] {request.method} {request.url}\n{tb}", flush=True)
    return JSONResponse(
        status_code=500,
        content={"detail": "Internal server error"},
    )

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True, allow_methods=["*"], allow_headers=["*"],
)

# GZip compression for API responses
from starlette.middleware.gzip import GZipMiddleware
app.add_middleware(GZipMiddleware, minimum_size=500)

# ── Auth middleware — protect all /api/* except public endpoints ──
from starlette.middleware.base import BaseHTTPMiddleware
from app.auth import decode_token

PUBLIC_PATHS = {
    "/api/auth/register", "/api/auth/login", "/api/auth/forgot-password",
    "/api/auth/reset-password", "/api/auth/check-username", "/api/auth/check-email",
    "/api/health",
}
PUBLIC_PREFIXES = ("/docs", "/openapi.json", "/redoc", "/api/tutorial/", "/api/sandbox/", "/api/bot/")

class AuthMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        path = request.url.path
        # Skip non-API and public paths
        if not path.startswith("/api/") or path in PUBLIC_PATHS or path.startswith(PUBLIC_PREFIXES):
            return await call_next(request)
        # Check Authorization header
        auth_header = request.headers.get("authorization", "")
        if not auth_header.startswith("Bearer "):
            return JSONResponse(status_code=401, content={"detail": "Not authenticated"})
        token = auth_header[7:]
        try:
            payload = decode_token(token)
            request.state.user_id = payload.get("sub", "")
            request.state.username = payload.get("username", "")
        except Exception:
            return JSONResponse(status_code=401, content={"detail": "Invalid or expired token"})
        return await call_next(request)

app.add_middleware(AuthMiddleware)

# ── Mount all routers ──
app.include_router(auth_router)
app.include_router(project_router)
app.include_router(overview_router)
app.include_router(diagnose_router)
app.include_router(design_router)
app.include_router(simulate_router)
app.include_router(mobilize_router)
app.include_router(export_router)
app.include_router(job_arch_router)
app.include_router(ja_mapping_router)
app.include_router(wd_router)
app.include_router(mgr_router)
app.include_router(scorecard_router)
app.include_router(diagnostic_router)
app.include_router(synthesis_router)
app.include_router(org_chart_router)
app.include_router(ai_router)
app.include_router(analytics_router)
app.include_router(tutorial_router)
app.include_router(export_ext_router)
if skills_router:
    app.include_router(skills_router)
if notes_router:
    app.include_router(notes_router)
if job_content_router:
    app.include_router(job_content_router)

# Skills Library
try:
    from app.routes_skills_library import router as skills_library_router
except ImportError:
    skills_library_router = None
if skills_library_router:
    app.include_router(skills_library_router)

# O*NET Skills Map Engine
try:
    from app.routes_onet import router as onet_router
except ImportError:
    onet_router = None
if onet_router:
    app.include_router(onet_router)

# Agent system
from app.routes_agents import router as agents_router
app.include_router(agents_router)

# Natural Language Query
from app.routes_nlq import router as nlq_router
app.include_router(nlq_router)

# Flight Recorder
from app.routes_flight_recorder import router as flight_recorder_router
app.include_router(flight_recorder_router)

# Decision persistence
from app.routes_decisions import router as decisions_router
app.include_router(decisions_router)

# Platform Concierge
try:
    from app.routes_concierge import router as concierge_router
except ImportError:
    concierge_router = None
if concierge_router:
    app.include_router(concierge_router)

# AI Provider abstraction
from app.ai_providers import call_ai_sync, get_ai_status

@app.post("/api/ai/ask")
def ask_ai_endpoint(request: dict):
    """Unified AI endpoint — routes all requests through Claude"""
    prompt = request.get("prompt", "")
    task_type = request.get("task_type", "general")
    system = request.get("system", None)
    if not prompt:
        return {"error": "No prompt provided"}
    try:
        response, provider = call_ai_sync(prompt, task_type, system)
        return {"response": response, "provider": provider}
    except Exception as e:
        return {"error": str(e)}

@app.get("/api/ai/providers")
def ai_providers_status():
    """Return status of all AI providers"""
    return get_ai_status()


# ── AI Intelligence Layer endpoints ──
from app.ai.insights import get_insights
from app.ai.recommendations import get_recommendations
from app.ai.query import parse_query

@app.post("/api/ai/insights")
def ai_insights_endpoint(request: Request, body: dict):
    """Get AI observations for a specific module."""
    module = body.get("module", "")
    data_summary = body.get("data_summary", "")
    context = body.get("context", "")
    filters = body.get("filters", {})
    project_id = body.get("project_id", "")
    if not module:
        return {"error": "module is required"}
    return get_insights(module, data_summary, context, filters, project_id)

@app.post("/api/ai/query")
def ai_query_endpoint(request: Request, body: dict):
    """Parse a natural language question and return answer + navigation."""
    question = body.get("question", "")
    data_summary = body.get("data_summary", "")
    context = body.get("context", "")
    if not question:
        return {"error": "question is required"}
    return parse_query(question, data_summary, context)

@app.get("/api/ai/recommendations")
def ai_recommendations_endpoint(request: Request):
    """Get contextual next-step recommendations."""
    # Query params
    from urllib.parse import unquote
    completed = request.query_params.get("completed", "")
    completed_list = [m.strip() for m in completed.split(",") if m.strip()]
    current = request.query_params.get("current", "")
    context = unquote(request.query_params.get("context", ""))
    has_workforce = request.query_params.get("has_workforce", "false") == "true"
    has_work_design = request.query_params.get("has_work_design", "false") == "true"
    has_skills = request.query_params.get("has_skills", "false") == "true"
    data_status = {"has_workforce": has_workforce, "has_work_design": has_work_design, "has_skills": has_skills}
    return get_recommendations(completed_list, data_status, current, context)

@app.post("/api/ai/chat")
def ai_chat_endpoint(request: Request, body: dict):
    """Copilot chat message — contextual conversation."""
    message = body.get("message", "")
    module = body.get("module", "")
    context = body.get("context", "")
    history = body.get("history", [])
    if not message:
        return {"error": "message is required"}
    # Format history for the prompt
    history_text = ""
    for h in history[-10:]:  # Last 10 messages
        role = h.get("role", "user")
        text = h.get("text", "")
        history_text += f"{'User' if role == 'user' else 'Assistant'}: {text}\n"
    from app.ai.prompts import CHAT_SYSTEM, CHAT_PROMPT
    from app.ai_providers import claude_available as _ai_ok
    if not _ai_ok:
        return {
            "response": "AI Copilot is available when ANTHROPIC_API_KEY is configured. Sandbox companies include pre-built insights — try exploring those modules directly.",
            "source": "fallback",
        }
    prompt = CHAT_PROMPT.format(module=module, context=context, history=history_text, message=message)
    try:
        from app.ai_providers import call_claude_sync as _call
        response = _call(prompt, system=CHAT_SYSTEM, model="claude-haiku-4-5-20251001", max_tokens=1024)
        return {"response": response, "source": "ai"}
    except Exception as e:
        return {"response": f"I couldn't process that right now: {str(e)}", "source": "error"}


# ── Bot Session Engine endpoints (no auth required) ──
from app.bot.engine import bot_engine

@app.post("/api/bot/start")
def bot_start(body: dict):
    project_id = body.get("project_id", "")
    model_id = body.get("model_id", "")
    mode = body.get("mode", "guided")
    if not model_id:
        return {"error": "model_id is required"}
    return bot_engine.start_session(project_id, model_id, mode)

@app.get("/api/bot/{session_id}/status")
def bot_status(session_id: str):
    return bot_engine.get_status(session_id)

@app.post("/api/bot/{session_id}/action/{action_name}")
def bot_run_action(session_id: str, action_name: str):
    return bot_engine.run_action(session_id, action_name)

@app.post("/api/bot/{session_id}/command")
def bot_command(session_id: str, body: dict):
    text = body.get("text", "")
    if not text:
        return {"error": "text is required"}
    return bot_engine.process_command(session_id, text)

@app.post("/api/bot/{session_id}/autopilot")
def bot_autopilot(session_id: str):
    return bot_engine.run_autopilot(session_id)

@app.get("/api/bot/{session_id}/findings")
def bot_findings(session_id: str):
    return bot_engine.get_findings(session_id)

@app.get("/api/bot/{session_id}/stream")
async def bot_stream(session_id: str):
    """SSE stream for real-time bot updates."""
    import asyncio
    import json as _json
    ctx = bot_engine.sessions.get(session_id)
    if not ctx:
        return JSONResponse(status_code=404, content={"error": "Session not found"})

    async def event_generator():
        last_log_idx = 0
        last_findings_count = 0
        last_status = ""
        last_progress = -1

        while True:
            ctx = bot_engine.sessions.get(session_id)
            if not ctx:
                yield f"event: error\ndata: {_json.dumps({'type': 'error', 'content': 'Session lost'})}\n\n"
                break

            # New log entries
            current_log_len = len(ctx.activity_log)
            if current_log_len > last_log_idx:
                for entry in ctx.activity_log[last_log_idx:current_log_len]:
                    etype = entry.get("type", "narration")
                    yield f"event: {etype}\ndata: {_json.dumps(entry, default=str)}\n\n"
                last_log_idx = current_log_len

            # New findings
            current_findings = len(ctx.findings)
            if current_findings > last_findings_count:
                for f in ctx.findings[last_findings_count:current_findings]:
                    yield f"event: finding\ndata: {_json.dumps(f, default=str)}\n\n"
                last_findings_count = current_findings

            # Status changes
            if ctx.status != last_status:
                last_status = ctx.status
                yield f"event: status\ndata: {_json.dumps({'type': 'status', 'status': ctx.status})}\n\n"

            # Progress changes
            prog = ctx.get_progress()
            if prog["completed"] != last_progress:
                last_progress = prog["completed"]
                yield f"event: progress\ndata: {_json.dumps({'type': 'progress', **prog})}\n\n"

            # Visualization data when action completes
            if ctx.completed_actions and len(ctx.completed_actions) > last_progress - 1:
                last_action = ctx.completed_actions[-1] if ctx.completed_actions else None
                if last_action and last_action in ctx.analysis_results:
                    try:
                        viz_data = _json.dumps({"type": "visualization", "action": last_action}, default=str)
                        yield f"event: visualization\ndata: {viz_data}\n\n"
                    except Exception:
                        pass

            if ctx.status in ("completed", "error"):
                yield f"event: done\ndata: {_json.dumps({'type': 'done', 'status': ctx.status})}\n\n"
                break

            await asyncio.sleep(0.5)

    return StreamingResponse(event_generator(), media_type="text/event-stream", headers={"Cache-Control": "no-cache", "Connection": "keep-alive", "X-Accel-Buffering": "no"})

@app.post("/api/bot/{session_id}/pause")
def bot_pause(session_id: str):
    return bot_engine.pause(session_id)

@app.post("/api/bot/{session_id}/resume")
def bot_resume(session_id: str):
    return bot_engine.resume(session_id)


@app.get("/api/health")
def health_check():
    """Quick health check — returns version info and DB status."""
    import sys
    db_ok = False
    db_error = None
    try:
        from app.auth import get_db, UserDB
        db = next(get_db())
        db.query(UserDB).count()
        db_ok = True
    except Exception as e:
        db_error = str(e)
    return {
        "status": "ok" if db_ok else "degraded",
        "python": sys.version,
        "db": "connected" if db_ok else f"ERROR: {db_error}",
        "db_type": "postgresql" if "postgresql" in (os.environ.get("DATABASE_URL", "") or "") else "sqlite",
    }

# ═══════════════════════════════════════════════════════════════════════
# MODELS / UPLOAD / RESET (kept in main.py — core app lifecycle)
# ═══════════════════════════════════════════════════════════════════════


def _user_id(request: Request) -> str:
    """Extract authenticated user_id from middleware-set state."""
    return getattr(request.state, "user_id", "")


def _user_models(request: Request) -> list[str]:
    """Return model IDs belonging to the current user."""
    uid = _user_id(request)
    if not uid:
        return store.get_model_ids()
    return [m for m in store.get_model_ids() if m.startswith(f"{uid}:") or not ":" in m]


@app.get("/api/models")
def list_models(request: Request):
    try:
        uid = _user_id(request)
        all_models = store.get_model_ids()
        # Return user-scoped models (prefixed with user_id:) + legacy unprefixed
        user_models = [m for m in all_models if m.startswith(f"{uid}:") or ":" not in m]
        # Strip user prefix for display
        display_models = [m.split(":", 1)[1] if m.startswith(f"{uid}:") else m for m in user_models]
        last = store.last_loaded_model_id or ""
        display_last = last.split(":", 1)[1] if last.startswith(f"{uid}:") else last
        return _safe({"models": display_models, "last_loaded": display_last})
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(500, detail=str(e))


@app.post("/api/upload/preview")
async def upload_preview(request: Request, files: list[UploadFile] = File(...)):
    """Preview uploaded file without ingesting. Returns sheet previews, column mapping suggestions, and validation results."""
    from app.schemas_definitions import SCHEMAS, COMMON_ALIASES
    from app.helpers import normalize_column_names
    from app.store import _load_excel_or_csv, classify_dataframe

    results = []

    for f in files:
        ext = Path(f.filename or "").suffix.lower()
        if ext not in ALLOWED_EXTENSIONS:
            results.append({"file": f.filename, "error": f"Unsupported file type: {ext}"})
            continue

        content = await f.read()
        size = len(content)

        if size > MAX_FILE_SIZE_BYTES:
            results.append({"file": f.filename, "error": f"File too large: {size / 1024 / 1024:.1f}MB (max {MAX_FILE_SIZE_MB}MB)"})
            continue

        # Load sheets
        try:
            sheets = _load_excel_or_csv(content, f.filename or "upload")
        except Exception as e:
            results.append({"file": f.filename, "error": f"Failed to parse: {str(e)[:200]}"})
            continue

        file_result = {
            "file": f.filename,
            "size_bytes": size,
            "size_label": f"{size / 1024:.0f}KB" if size < 1024 * 1024 else f"{size / 1024 / 1024:.1f}MB",
            "sheet_count": len(sheets),
            "sheets": [],
        }

        for sheet_name, df in sheets.items():
            if df.empty:
                continue

            # Auto-detect dataset type
            dtype = classify_dataframe(df, sheet_name)

            # Get schema for this type
            schema = SCHEMAS.get(dtype, {})
            expected_cols = schema.get("all", [])
            required_cols = schema.get("required", [])

            # Build column mapping suggestions
            normalized = normalize_column_names(df.copy())
            actual_cols = list(normalized.columns)

            mappings = []
            mapped_targets: set[str] = set()

            for col in actual_cols:
                col_str = str(col).strip()
                # Try exact match
                if col_str in expected_cols:
                    mappings.append({"source": col_str, "target": col_str, "confidence": 1.0, "status": "exact_match"})
                    mapped_targets.add(col_str)
                # Try alias match
                elif col_str in COMMON_ALIASES and COMMON_ALIASES[col_str] in expected_cols:
                    target = COMMON_ALIASES[col_str]
                    mappings.append({"source": col_str, "target": target, "confidence": 0.9, "status": "alias_match"})
                    mapped_targets.add(target)
                # Try case-insensitive match
                else:
                    match = None
                    for exp in expected_cols:
                        if col_str.lower().replace("_", " ").replace("-", " ") == exp.lower().replace("_", " ").replace("-", " "):
                            match = exp
                            break
                    if match and match not in mapped_targets:
                        mappings.append({"source": col_str, "target": match, "confidence": 0.85, "status": "fuzzy_match"})
                        mapped_targets.add(match)
                    else:
                        mappings.append({"source": col_str, "target": None, "confidence": 0, "status": "unmapped"})

            # Check for missing required columns
            missing_required = [c for c in required_cols if c not in mapped_targets]

            # Validation checks
            validation: dict = {
                "row_count": len(df),
                "column_count": len(df.columns),
                "errors": [],
                "warnings": [],
            }

            # Missing required columns -> errors
            for col in missing_required:
                validation["errors"].append({"type": "missing_required", "column": col, "message": f"Required column '{col}' is not mapped"})

            # Blank cells per mapped column
            for m in mappings:
                if m["target"] and m["source"] in df.columns:
                    blanks = int(df[m["source"]].isna().sum() + (df[m["source"]].astype(str).str.strip() == "").sum())
                    if blanks > 0:
                        pct = round(blanks / len(df) * 100, 1)
                        severity = "error" if m["target"] in required_cols and pct > 50 else "warning"
                        if severity == "error":
                            validation["errors"].append({"type": "blank_required", "column": m["target"], "count": blanks, "percentage": pct, "message": f"'{m['target']}' has {blanks} blank values ({pct}%)"})
                        else:
                            validation["warnings"].append({"type": "blank_values", "column": m["target"], "count": blanks, "percentage": pct, "message": f"'{m['target']}' has {blanks} blank values ({pct}%)"})

            # Duplicate IDs
            id_cols = [m["source"] for m in mappings if m["target"] in ("Employee ID", "Job Code", "Task ID") and m["source"] in df.columns]
            for id_col in id_cols:
                dupes = int(df[id_col].dropna().duplicated().sum())
                if dupes > 0:
                    validation["warnings"].append({"type": "duplicates", "column": id_col, "count": dupes, "message": f"'{id_col}' has {dupes} duplicate values"})

            # Orphan manager IDs
            emp_id_src = next((m["source"] for m in mappings if m["target"] == "Employee ID"), None)
            mgr_id_src = next((m["source"] for m in mappings if m["target"] == "Manager ID"), None)
            if emp_id_src and mgr_id_src and emp_id_src in df.columns and mgr_id_src in df.columns:
                emp_ids = set(df[emp_id_src].dropna().astype(str))
                mgr_ids = set(df[mgr_id_src].dropna().astype(str)) - {"", "nan", "None"}
                orphans = mgr_ids - emp_ids
                if orphans:
                    validation["warnings"].append({"type": "orphan_managers", "count": len(orphans), "message": f"{len(orphans)} Manager IDs reference employees not in the dataset"})

            # Preview rows (first 5)
            preview_rows = []
            for _, row in df.head(5).iterrows():
                preview_rows.append({str(c): str(v) if pd.notna(v) else "" for c, v in row.items()})

            # Detection confidence
            unmapped = len([m for m in mappings if m["status"] == "unmapped"])
            mapped = len([m for m in mappings if m["status"] != "unmapped"])
            detection_confidence = round(mapped / max(len(mappings), 1) * 100)

            sheet_result = {
                "sheet_name": sheet_name,
                "detected_type": dtype,
                "detection_confidence": detection_confidence,
                "detection_label": f"This looks like {dtype.replace('_', ' ').title()} data ({detection_confidence}% match)" if dtype != "unknown" else "Could not auto-detect dataset type",
                "row_count": len(df),
                "columns": actual_cols,
                "expected_columns": expected_cols,
                "required_columns": required_cols,
                "mappings": mappings,
                "missing_required": missing_required,
                "unmapped_count": unmapped,
                "preview": preview_rows,
                "validation": validation,
                "can_import": len(validation["errors"]) == 0,
            }
            file_result["sheets"].append(sheet_result)

        results.append(file_result)

    return _safe({"files": results})


@app.post("/api/upload/import")
async def upload_import(request: Request):
    """Import a previously previewed file with confirmed column mappings."""
    body = await request.json()
    # This endpoint accepts the file_id and mapping overrides
    # For now, delegate to the existing upload flow
    # The frontend will re-upload with mappings confirmed
    return _safe({"status": "ok"})


@app.get("/api/upload/validation/{model_id}")
async def get_validation_report(model_id: str, request: Request):
    """Get the validation report for the most recent upload."""
    from app.schemas_definitions import SCHEMAS
    uid = _user_id(request)
    model_id = store.resolve_model_id(model_id, user_id=uid)
    if model_id not in store.datasets:
        return _safe({"model_id": model_id, "status": "not_found"})

    bundle = store.datasets[model_id]
    report: dict = {"model_id": model_id, "datasets": {}}

    for dtype in ["workforce", "job_catalog", "work_design", "org_design", "operating_model", "change_management"]:
        df = bundle.get(dtype)
        if df is not None and not df.empty:
            schema = SCHEMAS.get(dtype, {})
            required = schema.get("required", [])

            issues = []
            for col in required:
                if col not in df.columns:
                    issues.append({"severity": "error", "message": f"Missing required column: {col}"})
                else:
                    blanks = int(df[col].isna().sum())
                    if blanks > 0:
                        issues.append({"severity": "warning", "message": f"{col}: {blanks} blank values ({round(blanks / len(df) * 100, 1)}%)"})

            report["datasets"][dtype] = {
                "rows": len(df),
                "columns": len(df.columns),
                "required_present": len([c for c in required if c in df.columns]),
                "required_total": len(required),
                "issues": issues,
                "status": "error" if any(i["severity"] == "error" for i in issues) else "warning" if issues else "clean",
            }

    return _safe(report)


@app.post("/api/upload")
async def upload_files(request: Request, files: list[UploadFile] = File(...)):
    uid = _user_id(request)
    # Validate files
    for f in files:
        ext = Path(f.filename).suffix.lower() if f.filename else ""
        if ext not in ALLOWED_EXTENSIONS:
            raise HTTPException(400, f"Invalid file type '{ext}'. Allowed: {', '.join(ALLOWED_EXTENSIONS)}")
    file_data = []
    stored_urls = []
    for f in files:
        content = await f.read()
        if len(content) > MAX_FILE_SIZE_BYTES:
            raise HTTPException(400, f"File '{f.filename}' exceeds {MAX_FILE_SIZE_MB}MB limit")
        file_data.append((f.filename, content))
        # Persist raw file to R2 or local storage for durability
        try:
            unique_name = make_unique_filename(f.filename or "upload", user_id=uid)
            url = file_storage.upload_file(content, unique_name)
            stored_urls.append(url)
        except Exception as e:
            print(f"[Storage] Failed to persist {f.filename}: {e}")
            stored_urls.append(None)

    s = store.process_uploads(file_data, user_id=uid)
    active_model = store.last_loaded_model_id or ""
    display_model = active_model.split(":", 1)[1] if active_model.startswith(f"{uid}:") else active_model
    user_models = [m for m in store.get_model_ids() if m.startswith(f"{uid}:") or ":" not in m]
    display_models = [m.split(":", 1)[1] if m.startswith(f"{uid}:") else m for m in user_models]

    # Flight recorder
    from app.flight_recorder import recorder
    filenames = [fd[0] for fd in file_data]
    recorder.record(display_model or "unknown", uid, getattr(request.state, "username", ""),
        "data.uploaded", "overview",
        f"Data uploaded: {', '.join(filenames)}",
        f"{len(s)} sheets loaded into model {display_model}",
        data_snapshot={"after": {"sheets": len(s), "model": display_model, "files": filenames, "storage_urls": [u for u in stored_urls if u]}},
        significance="major", tags=["upload"])

    return _safe({
        "sheets_loaded": len(s),
        "summary": _j(s),
        "active_model": display_model,
        "jobs": _job_list_for_model(store, active_model),
        "models": display_models,
    })


@app.post("/api/reset")
def reset_data(request: Request):
    uid = _user_id(request)
    # Only reset this user's data
    to_remove = [m for m in store.get_model_ids() if m.startswith(f"{uid}:")]
    for m in to_remove:
        if m in store.datasets:
            del store.datasets[m]
    return {"ok": True}


@app.get("/api/template")
def download_template():
    """Generate a multi-tab Excel template with schemas, example data, dropdowns, and Op Model guidance."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    NAVY, BLUE, TEAL, GREEN, AMBER = "0B1120", "3B82F6", "0891B2", "059669", "D97706"
    WHITE, LTBLUE, LGRAY, MGRAY, DKGRAY = "FFFFFF", "DBEAFE", "F1F5F9", "E2E8F0", "64748B"
    hdr_font = Font(name="Arial", bold=True, color=WHITE, size=10)
    hdr_fill = PatternFill("solid", fgColor=NAVY)
    req_fill = PatternFill("solid", fgColor=LTBLUE)
    opt_fill = PatternFill("solid", fgColor=LGRAY)
    hint_font = Font(name="Arial", italic=True, color=DKGRAY, size=9)
    ex_font = Font(name="Arial", color="8892A8", size=9)
    ex_fill = PatternFill("solid", fgColor="F1F5F9")
    title_font = Font(name="Arial", bold=True, size=14, color=NAVY)
    subtitle_font = Font(name="Arial", size=10, color=DKGRAY)
    guide_title = Font(name="Arial", bold=True, size=12, color=TEAL)
    guide_label = Font(name="Arial", bold=True, size=10, color=NAVY)
    guide_font = Font(name="Arial", size=9, color=DKGRAY)
    thin_border = Border(left=Side(style="thin", color=MGRAY), right=Side(style="thin", color=MGRAY),
                         top=Side(style="thin", color=MGRAY), bottom=Side(style="thin", color=MGRAY))
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap = Alignment(vertical="top", wrap_text=True)

    # ── Tab definitions: columns + example row ──
    TABS = {
        "1. Workforce": {
            "desc": "Employee roster — org hierarchy, comp, classification. Powers: Overview, Org Diagnostics, Filters. ★★★",
            "color": BLUE,
            "columns": [
                ("Model ID", True, "Your org/company name"), ("Employee ID", True, "Unique employee ID"), ("Employee Name", True, "Full name"),
                ("Manager ID", False, "Manager's Employee ID"), ("Manager Name", False, "Manager's name"), ("Function ID", False, "Function (Finance, HR...)"),
                ("Job Family", False, "Job family group"), ("Sub-Family", False, "Sub-function"), ("Geography", False, "Country/Region"),
                ("Career Track", False, "Executive/Manager/IC/Analyst"), ("Career Level", False, "Grade code (L2, L5...)"), ("Job Title", False, "Current job title"),
                ("Job Description", False, "Role description"), ("Department", False, "Department name"), ("Org Unit", False, "Business unit"),
                ("FTE", False, "1.0 = full-time"), ("Base Pay", False, "Annual base salary"), ("Total Cash", False, "Total cash comp"),
                ("Hire Date", False, "YYYY-MM-DD"), ("Performance Rating", False, "Rating"), ("Critical Role", False, "Yes/No"),
            ],
            "example": ["Acme_Corp", "E001", "Jane Smith", "", "", "Finance", "Corporate Functions", "FP&A", "US",
                        "Analyst", "L2", "Financial Analyst", "Supports budgeting and reporting", "Finance", "Corporate",
                        1.0, 95000, 102000, "2022-03-15", "Exceeds", "No"],
        },
        "2. Work Design": {
            "desc": "Task-level breakdown per job. Each job's tasks MUST sum to 100% time. Powers: Design Lab, AI Prioritization, Scenarios. ★★★",
            "color": AMBER,
            "columns": [
                ("Model ID", True, "Same Model ID"), ("Function ID", False, "Function"), ("Job Family", False, "Family"), ("Sub-Family", False, "Sub-function"),
                ("Geography", False, "Region"), ("Career Track", False, "Track"), ("Career Level", False, "Level"),
                ("Job Title", True, "Must match Workforce title"), ("Job Description", False, "Role summary"),
                ("Workstream", False, "Workflow group"), ("Task ID", False, "Unique ID (e.g., FA-T1)"), ("Task Name", True, "Descriptive task name"),
                ("Description", False, "What the task involves"), ("AI Impact", False, "High/Moderate/Low"), ("Est Hours/Week", False, "Hours per week"),
                ("Current Time Spent %", False, "% of role time (sum to 100)"), ("Time Saved %", False, "% AI can save"),
                ("Task Type", False, "Repetitive/Variable"), ("Interaction", False, "Independent/Interactive/Collaborative"),
                ("Logic", False, "Deterministic/Probabilistic/Judgment-heavy"), ("Primary Skill", False, "Main skill"), ("Secondary Skill", False, "Supporting skill"),
            ],
            "example": ["Acme_Corp", "Finance", "Corporate Functions", "FP&A", "US", "Analyst", "L2",
                        "Financial Analyst", "Supports budgeting", "Reporting", "FA-T1", "Refresh Management Report Pack",
                        "Update monthly board reports", "High", 4, 10, 5,
                        "Repetitive", "Independent", "Deterministic", "Reporting", "Excel"],
        },
        "3. Operating Model": {
            "desc": "Capability map — layers, scope, hierarchy, ownership. Powers: Structure, Maturity, Decisions, Workflow. ★★★ See guidance below row 7.",
            "color": TEAL,
            "columns": [
                ("Model ID", True, "Same Model ID"), ("Scope", True, "Enterprise or function name (Finance, HR, Technology...)"),
                ("Layer", True, "Governance / Core Components / Shared Services / Enabling / Interface"),
                ("Level 1", False, "Capability area (e.g., Planning)"), ("Level 2", False, "Sub-capability (e.g., Forecasting)"),
                ("Level 3", False, "Process/activity"), ("Level 4", False, "Granular detail"),
                ("Description", False, "Describe — mention 'shared/central' or 'embedded/dedicated' for service model detection"),
                ("Owner", False, "CRITICAL: Capability owner title — enables Decision Rights analysis"),
                ("Function ID", False, "Owning function"), ("Job Family", False, "Related family"), ("Sub-Family", False, "Sub-family"),
                ("Geography", False, "Global / US / UK / India — drives Scope Distribution"), ("Career Track", False, "Track"), ("Career Level", False, "Level"),
            ],
            "example": ["Acme_Corp", "Finance", "Core Components", "Planning", "Forecasting", "Monthly Forecast", "Revenue Build",
                        "Monthly revenue forecasting process — embedded in Finance team", "Director, FP&A",
                        "Finance", "Corporate Functions", "FP&A", "US", "", ""],
        },
        "4. Org Design": {
            "desc": "Org hierarchy for span of control, layer analysis. Can overlap with Workforce. ★★☆",
            "color": GREEN,
            "columns": [
                ("Model ID", True, "Same Model ID"), ("Employee ID", True, "Unique ID"), ("Employee Name", True, "Full name"),
                ("Manager ID", False, "Manager's Employee ID"), ("Manager Name", False, "Manager name"), ("Function ID", False, "Function"),
                ("Job Family", False, "Family"), ("Sub-Family", False, "Sub-function"), ("Geography", False, "Region"),
                ("Career Track", False, "Track"), ("Career Level", False, "Level"), ("Job Title", False, "Title"),
                ("Job Description", False, "Description"), ("Department", False, "Department"), ("Org Unit", False, "Org unit"),
            ],
            "example": ["Acme_Corp", "E001", "Jane Smith", "", "", "Finance", "Corporate Functions", "Leadership",
                        "US", "Executive", "L10", "Chief Financial Officer", "Leads all finance", "Finance", "Enterprise"],
        },
        "5. Job Catalog": {
            "desc": "Standardized job architecture — families, levels, career paths. Powers: Career Architecture, Role Clusters. ★★☆",
            "color": "8B5CF6",
            "columns": [
                ("Model ID", True, "Same Model ID"), ("Job Code", False, "Internal code"), ("Job Title", True, "Standardized title"),
                ("Standard Title", False, "Benchmark title"), ("Function ID", False, "Function"), ("Job Family", False, "Family"),
                ("Sub-Family", False, "Sub-function"), ("Geography", False, "Region"), ("Career Track", False, "Track"),
                ("Career Level", False, "Level"), ("Manager or IC", False, "Manager/IC"), ("Job Description", False, "Description"),
                ("Skills", False, "Semicolon-separated skills"), ("Role Purpose", False, "One-line purpose"),
            ],
            "example": ["Acme_Corp", "FIN-ANL-L2", "Financial Analyst", "Financial Analyst", "Finance",
                        "Corporate Functions", "FP&A", "US", "Analyst", "L2", "IC",
                        "Supports budgeting and reporting", "Excel;Analysis;Reporting;PowerBI", "Provide financial insight and support."],
        },
        "6. Change Management": {
            "desc": "Transformation initiatives, RACI, roadmap milestones. Powers: Roadmap, Risk Analysis, Mobilize. ★★☆",
            "color": "E11D48",
            "columns": [
                ("Model ID", True, "Same Model ID"), ("Function ID", False, "Function"), ("Job Family", False, "Family"),
                ("Sub-Family", False, "Sub-function"), ("Geography", False, "Region"), ("Career Track", False, "Track"), ("Career Level", False, "Level"),
                ("Job Title", False, "Affected title"), ("Task Name", False, "Task being changed"), ("Responsible", False, "RACI: does the work"),
                ("Accountable", False, "RACI: approves"), ("Consulted", False, "RACI: consulted"), ("Informed", False, "RACI: informed"),
                ("Initiative", False, "Initiative name"), ("Owner", False, "Initiative owner"), ("Priority", False, "High/Medium/Low"),
                ("Status", False, "Planned/In Progress/Complete"), ("Wave", False, "Wave 1/Wave 2/etc"), ("Start", False, "YYYY-MM-DD"),
                ("End", False, "YYYY-MM-DD"), ("Milestone", False, "Milestone name"), ("Date", False, "Milestone date"),
                ("Risk", False, "Risk description"), ("Dependency", False, "Dependencies"), ("Notes", False, "Additional notes"),
            ],
            "example": ["Acme_Corp", "Finance", "Corporate Functions", "FP&A", "US", "Analyst", "L2",
                        "Financial Analyst", "Reporting Automation", "Analyst", "Director", "Finance Lead", "CFO",
                        "Reporting Automation", "CFO", "High", "Planned", "Wave 1", "2026-01-01", "2026-03-31",
                        "Kickoff", "2026-01-01", "Resistance to change", "IT resource availability", "Phase 1 pilot"],
        },
    }

    wb = Workbook()

    # ── Instructions tab ──
    ws = wb.active
    ws.title = "Instructions"
    ws.sheet_properties.tabColor = BLUE
    ws.column_dimensions["A"].width = 90
    ws.column_dimensions["B"].width = 40
    lines = [
        ("AI Transformation Platform — Data Template", title_font),
        ("", subtitle_font),
        ("This workbook has 8 tabs — one for each dataset the platform ingests.", guide_font),
        ("Fill as many as you can. The more data, the richer the analysis.", guide_font),
        ("", guide_font),
        ("HOW TO USE", guide_title),
        ("1. Use a consistent Model ID across ALL tabs (e.g., your company name).", guide_font),
        ("2. Row 2 = column headers. Row 3 = REQUIRED/Optional. Row 4 = field hints.", guide_font),
        ("3. Row 5 = example data (delete before uploading).", guide_font),
        ("4. Fill your data starting from Row 6.", guide_font),
        ("5. Blue header columns are REQUIRED. Dark columns are optional.", guide_font),
        ("6. Job Titles must match across Workforce, Work Design, and Job Catalog.", guide_font),
        ("7. Work Design tasks per job MUST sum to 100% Current Time Spent %.", guide_font),
        ("8. Save and upload via the Upload button in the platform.", guide_font),
        ("", guide_font),
        ("TAB PRIORITY", guide_title),
        ("★★★ Workforce, Work Design, Operating Model — critical for core analysis", guide_font),
        ("★★☆ Org Design, Job Catalog, Change Management — important for depth", guide_font),
        ("", guide_font),
        ("OPERATING MODEL — SPECIAL GUIDANCE", guide_title),
        ("The Operating Model tab is the most nuanced dataset. Key rules:", guide_font),
        ("• Each ROW = one capability, process, or organizational component.", guide_font),
        ("• SCOPE = which domain this belongs to (Enterprise, Finance, HR, Technology, Legal...).", guide_font),
        ("• LAYER = architectural tier: Governance, Core Components, Shared Services, Enabling, Interface.", guide_font),
        ("• LEVELS 1-4 = hierarchical breakdown (L1=Planning > L2=Forecasting > L3=Monthly Forecast).", guide_font),
        ("• OWNER = who owns the capability — CRITICAL for Decision Rights dashboard.", guide_font),
        ("• DESCRIPTION = mention 'shared' or 'embedded' so the platform can detect Service Model.", guide_font),
        ("• GEOGRAPHY = use 'Global' for enterprise-wide items. Drives Scope Distribution chart.", guide_font),
        ("• Aim for 15-30 rows per function and 3-5 Governance rows for meaningful analysis.", guide_font),
        ("• See the Operating Model tab (row 8+) for a detailed field-by-field guide.", guide_font),
    ]
    for ri, (text, font) in enumerate(lines, 1):
        c = ws.cell(row=ri, column=1, value=text)
        c.font = font

    # ── Data tabs ──
    for tab_name, td in TABS.items():
        cols = td["columns"]
        ex = td.get("example", [])
        ws = wb.create_sheet(title=tab_name)
        ws.sheet_properties.tabColor = td["color"]
        ws.merge_cells(f"A1:{get_column_letter(len(cols))}1")
        ws["A1"] = td["desc"]
        ws["A1"].font = subtitle_font
        ws["A1"].alignment = wrap
        ws.row_dimensions[1].height = 30
        # Row 2: headers
        for ci, (cn, req, _hint) in enumerate(cols, 1):
            c = ws.cell(row=2, column=ci, value=cn)
            c.font = hdr_font
            c.fill = PatternFill("solid", fgColor=BLUE) if req else hdr_fill
            c.alignment = center
            c.border = thin_border
        # Row 3: required/optional
        for ci, (cn, req, _) in enumerate(cols, 1):
            c = ws.cell(row=3, column=ci, value="REQUIRED" if req else "Optional")
            c.font = Font(name="Arial", bold=req, size=9, color=BLUE if req else DKGRAY)
            c.fill = req_fill if req else opt_fill
            c.alignment = center
            c.border = thin_border
        # Row 4: hints
        for ci, (_, _, hint) in enumerate(cols, 1):
            c = ws.cell(row=4, column=ci, value=hint)
            c.font = hint_font
            c.alignment = wrap
            c.border = thin_border
        ws.row_dimensions[4].height = 40
        # Row 5: example data
        for ci, val in enumerate(ex, 1):
            c = ws.cell(row=5, column=ci, value=val)
            c.font = ex_font
            c.fill = ex_fill
            c.border = thin_border
        # Empty rows 6-30
        for ri in range(6, 31):
            for ci in range(1, len(cols) + 1):
                c = ws.cell(row=ri, column=ci, value="")
                c.border = thin_border
                if ri % 2 == 0:
                    c.fill = PatternFill("solid", fgColor="F8FAFC")
        # Column widths
        for ci, (cn, _, _) in enumerate(cols, 1):
            ws.column_dimensions[get_column_letter(ci)].width = max(len(cn) + 5, 15)
        ws.freeze_panes = "A6"
        ws.auto_filter.ref = f"A2:{get_column_letter(len(cols))}2"
        # Dropdowns
        for ci, (cn, _, _) in enumerate(cols, 1):
            cl = get_column_letter(ci)
            dv = None
            if cn == "AI Impact": dv = DataValidation(type="list", formula1='"High,Moderate,Low"', allow_blank=True)
            elif cn == "Task Type": dv = DataValidation(type="list", formula1='"Repetitive,Variable"', allow_blank=True)
            elif cn == "Interaction": dv = DataValidation(type="list", formula1='"Independent,Interactive,Collaborative"', allow_blank=True)
            elif cn == "Logic": dv = DataValidation(type="list", formula1='"Deterministic,Probabilistic,Judgment-heavy"', allow_blank=True)
            elif cn == "Priority": dv = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=True)
            elif cn == "Status": dv = DataValidation(type="list", formula1='"Planned,In Progress,Complete,On Hold"', allow_blank=True)
            elif cn == "Wave": dv = DataValidation(type="list", formula1='"Wave 1,Wave 2,Wave 3,Wave 4"', allow_blank=True)
            elif cn == "Manager or IC": dv = DataValidation(type="list", formula1='"Manager,IC"', allow_blank=True)
            elif cn == "Critical Role": dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
            elif cn == "Layer": dv = DataValidation(type="list", formula1='"Governance,Core Components,Shared Services,Enabling,Interface"', allow_blank=True)
            elif cn == "Performance Rating": dv = DataValidation(type="list", formula1='"Exceeds,Meets,Below,1,2,3,4,5"', allow_blank=True)
            if dv:
                dv.error = "Please select from the dropdown"
                ws.add_data_validation(dv)
                dv.add(f"{cl}6:{cl}500")

    # ── Operating Model tab: append guidance section below example ──
    ws_op = wb["3. Operating Model"]
    gr = 8  # start guidance after a gap row
    ws_op.cell(row=gr, column=1, value="OPERATING MODEL — FIELD-BY-FIELD GUIDE").font = guide_title
    ws_op.merge_cells(f"A{gr}:H{gr}")
    guidance = [
        ("Scope", "The domain this capability belongs to. Use 'Enterprise' for firm-wide governance.\nExamples: Enterprise, Finance, HR, Technology, Legal, Investments, Operations, Marketing."),
        ("Layer", "The architectural tier.\n• Governance = oversight bodies, committees, decision forums.\n• Core Components = primary capabilities that deliver the function's mission.\n• Shared Services = capabilities shared across business lines.\n• Enabling = infrastructure, platforms, tools.\n• Interface = touchpoints with the rest of the org (portals, dashboards)."),
        ("Level 1-4", "Hierarchical decomposition of the capability.\nL1 = Capability area (e.g., 'Planning').\nL2 = Sub-capability (e.g., 'Forecasting').\nL3 = Process (e.g., 'Monthly Revenue Forecast').\nL4 = Activity (e.g., 'Collect actuals from GL').\nNot all levels are required — use as many as your detail allows."),
        ("Description", "Free text describing the capability. IMPORTANT: If you mention 'shared', 'central', 'centralized', 'COE', or 'center of excellence', the platform tags it as Shared Service. If you mention 'embedded', 'dedicated', or 'local', it tags as Embedded. This drives the Service Model analysis."),
        ("Owner", "The role title who owns this capability. CRITICAL for Decision Rights.\nExamples: 'CFO', 'Director FP&A', 'CHRO', 'VP Engineering', 'Head of TA'.\nWithout Owner, the Decision Rights tab will show 'Unassigned'."),
        ("Geography", "Use 'Global' for enterprise-wide capabilities, or a specific region (US, UK, India, APAC) for local ones.\nDrives the Scope Distribution analysis (Global vs Local footprint)."),
        ("Best Practices", "• Aim for 15-30 rows per function for meaningful maturity and structure analysis.\n• Include 3-5 Governance rows (committees, forums) to power Decision Rights.\n• Mix Shared Services and Core Components rows for service model split analysis.\n• Fill Owner for every row — it's the #1 field that unlocks the Decisions dashboard.\n• Use consistent Function ID values that match your Workforce and Work Design tabs."),
    ]
    for i, (label, desc) in enumerate(guidance):
        r = gr + 1 + i
        ws_op.cell(row=r, column=1, value=label).font = guide_label
        ws_op.cell(row=r, column=2, value=desc).font = guide_font
        ws_op.cell(row=r, column=2).alignment = wrap
        ws_op.merge_cells(start_row=r, start_column=2, end_row=r, end_column=10)
        ws_op.row_dimensions[r].height = 60

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=AI_Transformation_Template.xlsx"})


# ── Per-module template download ──
MODULE_TAB_MAP = {
    "snapshot": "1. Workforce",
    "jobs": "5. Job Catalog",
    "scan": "2. Work Design",
    "design": "2. Work Design",
    "simulate": "2. Work Design",
    "build": "4. Org Design",
    "plan": "6. Change Management",
    "opmodel": "3. Operating Model",
}

@app.get("/api/template/{module_id}")
def download_module_template(module_id: str):
    """Download a single-tab Excel template for a specific module."""
    tab_name = MODULE_TAB_MAP.get(module_id)
    if not tab_name:
        raise HTTPException(status_code=404, detail=f"No template for module: {module_id}")

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    NAVY, BLUE, TEAL, GREEN, AMBER = "0B1120", "3B82F6", "0891B2", "059669", "D97706"
    WHITE, LTBLUE, LGRAY, MGRAY, DKGRAY = "FFFFFF", "DBEAFE", "F1F5F9", "E2E8F0", "64748B"
    hdr_font = Font(name="Arial", bold=True, color=WHITE, size=10)
    hdr_fill = PatternFill("solid", fgColor=NAVY)
    req_fill = PatternFill("solid", fgColor=LTBLUE)
    opt_fill = PatternFill("solid", fgColor=LGRAY)
    hint_font = Font(name="Arial", italic=True, color=DKGRAY, size=9)
    ex_font = Font(name="Arial", color="8892A8", size=9)
    ex_fill = PatternFill("solid", fgColor="F1F5F9")
    subtitle_font = Font(name="Arial", size=10, color=DKGRAY)
    thin_border = Border(left=Side(style="thin", color=MGRAY), right=Side(style="thin", color=MGRAY),
                         top=Side(style="thin", color=MGRAY), bottom=Side(style="thin", color=MGRAY))
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap = Alignment(vertical="top", wrap_text=True)

    # Reuse same tab definitions
    TABS = {
        "1. Workforce": {"desc":"Employee roster — powers Workforce Snapshot, Org Diagnostics, Filters.","color":BLUE,"columns":[("Model ID",True,"Your org/company name"),("Employee ID",True,"Unique employee ID"),("Employee Name",True,"Full name"),("Manager ID",False,"Manager's Employee ID"),("Manager Name",False,"Manager's name"),("Function ID",False,"Function (Finance, HR...)"),("Job Family",False,"Job family group"),("Sub-Family",False,"Sub-function"),("Geography",False,"Country/Region"),("Career Track",False,"Executive/Manager/IC/Analyst"),("Career Level",False,"Grade code"),("Job Title",False,"Current job title"),("Job Description",False,"Role description"),("Department",False,"Department name"),("Org Unit",False,"Business unit"),("FTE",False,"1.0 = full-time"),("Base Pay",False,"Annual base salary"),("Total Cash",False,"Total cash comp"),("Hire Date",False,"YYYY-MM-DD"),("Performance Rating",False,"Rating"),("Critical Role",False,"Yes/No")],"example":["Acme_Corp","E001","Jane Smith","","","Finance","Corporate Functions","FP&A","US","Analyst","L2","Financial Analyst","Supports budgeting","Finance","Corporate",1.0,95000,102000,"2022-03-15","Exceeds","No"]},
        "2. Work Design": {"desc":"Task-level breakdown per job. Each job's tasks MUST sum to 100% time. Powers: Design Lab, AI Scan, Simulator.","color":AMBER,"columns":[("Model ID",True,"Same Model ID"),("Function ID",False,"Function"),("Job Family",False,"Family"),("Sub-Family",False,"Sub-function"),("Geography",False,"Region"),("Career Track",False,"Track"),("Career Level",False,"Level"),("Job Title",True,"Must match Workforce title"),("Job Description",False,"Role summary"),("Workstream",False,"Workflow group"),("Task ID",False,"Unique ID"),("Task Name",True,"Descriptive task name"),("Description",False,"What the task involves"),("AI Impact",False,"High/Moderate/Low"),("Est Hours/Week",False,"Hours per week"),("Current Time Spent %",False,"% of role time (sum to 100)"),("Time Saved %",False,"% AI can save"),("Task Type",False,"Repetitive/Variable"),("Interaction",False,"Independent/Interactive/Collaborative"),("Logic",False,"Deterministic/Probabilistic/Judgment-heavy"),("Primary Skill",False,"Main skill"),("Secondary Skill",False,"Supporting skill")],"example":["Acme_Corp","Finance","Corporate Functions","FP&A","US","Analyst","L2","Financial Analyst","Supports budgeting","Reporting","FA-T1","Refresh Management Report Pack","Update monthly board reports","High",4,10,5,"Repetitive","Independent","Deterministic","Reporting","Excel"]},
        "3. Operating Model": {"desc":"Capability map — layers, scope, hierarchy, ownership. Powers: Operating Model Lab.","color":TEAL,"columns":[("Model ID",True,"Same Model ID"),("Scope",True,"Enterprise or function name"),("Layer",True,"Governance / Core Components / Shared Services / Enabling / Interface"),("Level 1",False,"Capability area"),("Level 2",False,"Sub-capability"),("Level 3",False,"Process/activity"),("Level 4",False,"Granular detail"),("Description",False,"Describe capability"),("Owner",False,"Capability owner title"),("Function ID",False,"Owning function"),("Job Family",False,"Related family"),("Sub-Family",False,"Sub-family"),("Geography",False,"Global/US/UK/India"),("Career Track",False,"Track"),("Career Level",False,"Level")],"example":["Acme_Corp","Finance","Core Components","Planning","Forecasting","Monthly Forecast","Revenue Build","Monthly revenue forecasting process","Director, FP&A","Finance","Corporate Functions","FP&A","US","",""]},
        "4. Org Design": {"desc":"Org hierarchy for span of control, layer analysis.","color":GREEN,"columns":[("Model ID",True,"Same Model ID"),("Employee ID",True,"Unique ID"),("Employee Name",True,"Full name"),("Manager ID",False,"Manager's Employee ID"),("Manager Name",False,"Manager name"),("Function ID",False,"Function"),("Job Family",False,"Family"),("Sub-Family",False,"Sub-function"),("Geography",False,"Region"),("Career Track",False,"Track"),("Career Level",False,"Level"),("Job Title",False,"Title"),("Job Description",False,"Description"),("Department",False,"Department"),("Org Unit",False,"Org unit")],"example":["Acme_Corp","E001","Jane Smith","","","Finance","Corporate Functions","Leadership","US","Executive","L10","Chief Financial Officer","Leads all finance","Finance","Enterprise"]},
        "5. Job Catalog": {"desc":"Standardized job architecture — families, levels, career paths.","color":"8B5CF6","columns":[("Model ID",True,"Same Model ID"),("Job Code",False,"Internal code"),("Job Title",True,"Standardized title"),("Standard Title",False,"Benchmark title"),("Function ID",False,"Function"),("Job Family",False,"Family"),("Sub-Family",False,"Sub-function"),("Geography",False,"Region"),("Career Track",False,"Track"),("Career Level",False,"Level"),("Manager or IC",False,"Manager/IC"),("Job Description",False,"Description"),("Skills",False,"Semicolon-separated skills"),("Role Purpose",False,"One-line purpose")],"example":["Acme_Corp","FIN-ANL-L2","Financial Analyst","Financial Analyst","Finance","Corporate Functions","FP&A","US","Analyst","L2","IC","Supports budgeting","Excel;Analysis;Reporting","Provide financial insight."]},
        "6. Change Management": {"desc":"Transformation initiatives, RACI, roadmap milestones. Powers: Change Planner.","color":"E11D48","columns":[("Model ID",True,"Same Model ID"),("Function ID",False,"Function"),("Job Family",False,"Family"),("Sub-Family",False,"Sub-function"),("Geography",False,"Region"),("Career Track",False,"Track"),("Career Level",False,"Level"),("Job Title",False,"Affected title"),("Task Name",False,"Task being changed"),("Initiative",False,"Initiative name"),("Owner",False,"Initiative owner"),("Priority",False,"High/Medium/Low"),("Status",False,"Planned/In Progress/Complete"),("Wave",False,"Wave 1/Wave 2/etc"),("Start",False,"YYYY-MM-DD"),("End",False,"YYYY-MM-DD"),("Risk",False,"Risk description"),("Notes",False,"Additional notes")],"example":["Acme_Corp","Finance","Corporate Functions","FP&A","US","Analyst","L2","Financial Analyst","Reporting Automation","Reporting Automation","CFO","High","Planned","Wave 1","2026-01-01","2026-03-31","Resistance to change","Phase 1 pilot"]},
    }

    td = TABS.get(tab_name)
    if not td:
        raise HTTPException(status_code=404, detail=f"Tab not found: {tab_name}")

    wb = Workbook()
    ws = wb.active
    ws.title = tab_name
    ws.sheet_properties.tabColor = td["color"]
    cols = td["columns"]
    ex = td.get("example", [])

    ws.merge_cells(f"A1:{get_column_letter(len(cols))}1")
    ws["A1"] = td["desc"]
    ws["A1"].font = subtitle_font
    ws["A1"].alignment = wrap
    ws.row_dimensions[1].height = 30
    for ci, (cn, req, _) in enumerate(cols, 1):
        c = ws.cell(row=2, column=ci, value=cn)
        c.font = hdr_font; c.fill = PatternFill("solid", fgColor=BLUE) if req else hdr_fill; c.alignment = center; c.border = thin_border
    for ci, (_, req, _) in enumerate(cols, 1):
        c = ws.cell(row=3, column=ci, value="REQUIRED" if req else "Optional")
        c.font = Font(name="Arial", bold=req, size=9, color=BLUE if req else DKGRAY); c.fill = req_fill if req else opt_fill; c.alignment = center; c.border = thin_border
    for ci, (_, _, hint) in enumerate(cols, 1):
        c = ws.cell(row=4, column=ci, value=hint); c.font = hint_font; c.alignment = wrap; c.border = thin_border
    ws.row_dimensions[4].height = 40
    for ci, val in enumerate(ex, 1):
        c = ws.cell(row=5, column=ci, value=val); c.font = ex_font; c.fill = ex_fill; c.border = thin_border
    for ri in range(6, 51):
        for ci in range(1, len(cols) + 1):
            c = ws.cell(row=ri, column=ci, value=""); c.border = thin_border
            if ri % 2 == 0: c.fill = PatternFill("solid", fgColor="F8FAFC")
    for ci, (cn, _, _) in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = max(len(cn) + 5, 15)
    ws.freeze_panes = "A6"
    for ci, (cn, _, _) in enumerate(cols, 1):
        cl = get_column_letter(ci)
        dv = None
        if cn == "AI Impact": dv = DataValidation(type="list", formula1='"High,Moderate,Low"', allow_blank=True)
        elif cn == "Task Type": dv = DataValidation(type="list", formula1='"Repetitive,Variable"', allow_blank=True)
        elif cn == "Interaction": dv = DataValidation(type="list", formula1='"Independent,Interactive,Collaborative"', allow_blank=True)
        elif cn == "Logic": dv = DataValidation(type="list", formula1='"Deterministic,Probabilistic,Judgment-heavy"', allow_blank=True)
        elif cn == "Priority": dv = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=True)
        elif cn == "Status": dv = DataValidation(type="list", formula1='"Planned,In Progress,Complete,On Hold"', allow_blank=True)
        elif cn == "Wave": dv = DataValidation(type="list", formula1='"Wave 1,Wave 2,Wave 3,Wave 4"', allow_blank=True)
        elif cn == "Layer": dv = DataValidation(type="list", formula1='"Governance,Core Components,Shared Services,Enabling,Interface"', allow_blank=True)
        if dv:
            dv.error = "Please select from the dropdown"; ws.add_data_validation(dv); dv.add(f"{cl}6:{cl}500")

    safe_name = tab_name.replace(". ", "_").replace(" ", "_")
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={safe_name}_Template.xlsx"})


# ── Domain routes (filter-options, job-options, overview, diagnose/*,
# ── design/*, simulate/*, mobilize/*, operating-model, export/datasets,
# ── export/download) are in their respective routes_*.py files.

@app.get("/")
def root():
    return {"ok": True, "app": "ai-transformation-platform", "version": "4.0"}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("app.main:app", host="0.0.0.0", port=8000, reload=True)
